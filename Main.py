import os
import sys
import signal
import time
import logging
import pandas as pd
from openpyxl import load_workbook  # used to open an existing  excel file

from selenium import webdriver
from selenium.webdriver.chrome.options import Options   # configure  the chrome options
from selenium.webdriver.chrome.service import Service   # manage the chrome services 
from webdriver_manager.chrome import ChromeDriverManager # Automatically manage chrome driver

from selenium.webdriver.common.by import By # Locate the Html tags
from selenium.webdriver.common.action_chains import ActionChains    # stimulate the mouse action
from selenium.webdriver.support.ui import WebDriverWait # waits for elements
from selenium.webdriver.support import expected_conditions as EC    # define the wait condition

# -------------------- Logging --------------------
def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler("scraping_log.txt"),
            logging.StreamHandler()
        ]
    )

# -------------------- Graceful Exit --------------------
def setup_interrupt(driver):        # managing interrupts
    def handle_interrupt(signum, frame):
        logging.warning("Scraping interrupted by user (Ctrl+C). Shutting down gracefully.")
        try:
            driver.quit()
        except:
            pass
        sys.exit()
    signal.signal(signal.SIGINT, handle_interrupt)

# -------------------- Browser Configuration --------------------
def init_browser():
    options = Options()
    options.add_argument("--headless")          # Run without opening broweser
    options.add_argument("--disable-gpu")       # No gpu required
    options.add_argument("--window-size=1920x1080") # window size
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

# -------------------- Excel Handling --------------------
def load_existing_data(excel_file):
    processed_bid_ids = set()               # handles duplicate values
    if os.path.exists(excel_file):
        df_existing = pd.read_excel(excel_file)
        logging.info(f"Existing Excel columns: {df_existing.columns.tolist()}")
        if "Bid ID" in df_existing.columns:
            processed_bid_ids = set(df_existing["Bid ID"].astype(str))
            logging.info(f"Loaded {len(processed_bid_ids)} previously scraped bids.")
        else:
            logging.warning("Column 'Bid ID' not found. Proceeding with empty ID set.")
    return processed_bid_ids

def save_bid_to_excel(df_row, excel_file):
    if os.path.exists(excel_file):
        try:
            book = load_workbook(excel_file)
            sheet = book.active
            current_row = sheet.max_row
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_row.to_excel(writer, index=False, header=False, startrow=current_row)
        except PermissionError as pe:
            logging.error(f"Excel file locked: {pe}")
    else:
        df_row.to_excel(excel_file, index=False)
        logging.info("Created new Excel file with headers.")

# -------------------- Scraping --------------------
def extract_modal_data(modal):
    contact_email = "N/A"
    ad_date = "N/A"
    deadline_response = "N/A"
    specific_message = "N/A"
    document_dict = {}

    try:
        contact_email = modal.find_element(By.XPATH, ".//a[contains(@href, 'mailto')]").text.strip()
    except:
        contact_email = "N/A"

    try:
        ad_date = modal.find_element(By.XPATH, ".//label[preceding-sibling::label[contains(text(),'Solicitation Ad Date')]]").text.strip()
    except:
        ad_date = "N/A"

    try:
        deadline_response = modal.find_element(By.XPATH, ".//label[preceding-sibling::label[contains(text(),'Deadline for Bid Responses')]]").text.strip()
    except:
        deadline_response = "N/A"

    try:
        specific_message = modal.find_element(By.XPATH, ".//h6[contains(@class, 'text-danger')]").text.strip()
    except:
        specific_message = "N/A"

    try:
        doc_links = modal.find_elements(By.XPATH, ".//div[@id='bidDocuments']//a")
        for link in doc_links:
            try:
                name = link.text.strip()
                href = link.get_attribute("href")
                document_dict[name] = href
            except:
                document_dict[name] = "N/A"
    except:
        document_dict = {}

    return contact_email, ad_date, deadline_response, specific_message, document_dict

def scrape_bid_rows(driver, wait, category_name, processed_bid_ids, excel_file):
    while True:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#jqGridBids tbody tr")))
        rows = driver.find_elements(By.CSS_SELECTOR, "#jqGridBids tbody tr")

        for row in rows:
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 7:
                    continue

                bid_id = cells[0].get_attribute("title")
                if str(bid_id) in processed_bid_ids:
                    logging.info(f"Skipping already scraped Bid ID: {bid_id}")
                    continue

                contract_number = cells[1].text.strip()
                title_link = cells[2].find_element(By.TAG_NAME, "a")
                title = title_link.text.strip()
                open_date = cells[3].text.strip()
                deadline = cells[4].text.strip()
                agency = cells[5].text.strip()
                unspsc = cells[6].text.strip()

                driver.execute_script("arguments[0].click();", title_link)
                time.sleep(2)

                modal = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "dynamicDialogInnerHtml"))
                )

                contact_email, ad_date, deadline_response, specific_message, document_dict = extract_modal_data(modal)

                ActionChains(driver).send_keys(u'\ue00c').perform()
                time.sleep(1)

                bid_data = {
                    "Category": category_name,
                    "Bid ID": bid_id,
                    "Contract Number": contract_number,
                    "Title": title,
                    "Open Date": open_date,
                    "Deadline": deadline,
                    "Agency": agency,
                    "UNSPSC": unspsc,
                    "Solicitation Ad Date": ad_date,
                    "Deadline for Bid Responses": deadline_response,
                    "Contact Email": contact_email,
                    "Important Message": specific_message,
                    "Documents": str(document_dict)
                }

                processed_bid_ids.add(str(bid_id))
                logging.info(f"Scraped Bid ID: {bid_id}")

                df_row = pd.DataFrame([bid_data])
                save_bid_to_excel(df_row, excel_file)
                logging.info("Saved bid to Excel immediately.")

            except Exception as e:
                logging.error(f"Error processing row: {e}")
                continue

        try:
            next_btn = driver.find_element(By.ID, "next_jqg1")
            next_class = next_btn.get_attribute("class")
            if "disabled" in next_class or "ui-jqgrid-disablePointerEvents" in next_class:
                break
            driver.execute_script("arguments[0].click();", next_btn)
            time.sleep(2)
        except Exception as e:
            logging.error(f"Pagination failed or ended: {e}")
            break

# -------------------- Run Scraper --------------------
def run_scraper():
    setup_logging()
    excel_file = "delaware_bids_all_categories.xlsx"
    driver = init_browser()
    setup_interrupt(driver)

    processed_bid_ids = load_existing_data(excel_file)
    driver.get("https://mmp.delaware.gov/Bids/")
    wait = WebDriverWait(driver, 15)

    categories = [("Open", "btnOpen")]
    for category_name, tab_id in categories:
        logging.info(f"Scraping category: {category_name}")
        wait.until(EC.element_to_be_clickable((By.ID, tab_id))).click()
        time.sleep(2)
        scrape_bid_rows(driver, wait, category_name, processed_bid_ids, excel_file)

    driver.quit()
    logging.info("Scraping complete.")

# -------------------- Entry Point --------------------
if __name__ == "__main__":
    run_scraper()
